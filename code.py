import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
import re

import re

def scrape_contact_data(url):
    # Function definition for scraping contact data
    print("Scraping contact data from:", url)
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')
    
    # Find all text content within the webpage
    all_text = soup.get_text(separator="\n", strip=True)
    
    # Initialize lists to store data
    company_names = []
    street_numbers = []
    postal_codes = []
    countries = []
    email_addresses = []
    phone_numbers = []
    fax_numbers = []
    websites = []
    
    # Search for patterns indicative of contact information
    # Email addresses pattern
    email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
    email_addresses.extend(re.findall(email_pattern, all_text))
    print("Number of email addresses found:", len(email_addresses))
    
    # Phone numbers pattern
    phone_pattern = r'(?:(?:\+|0{0,2})\d{1,4}[\s.-]?)?(?:\(\d{1,5}\)[\s.-]?)?\d{3,5}[\s.-]?\d{3,4}[\s.-]?\d{3,4}'
    phone_numbers.extend(re.findall(phone_pattern, all_text))
    print("Number of phone numbers found:", len(phone_numbers))
    
    # Fax numbers pattern
    fax_pattern = r'Fax:\s*([\d\s/-]+)'
    fax_numbers.extend(re.findall(fax_pattern, all_text))
    print("Number of fax numbers found:", len(fax_numbers))
    
    # Address pattern (assuming typical address format)
    address_pattern = r'\b\d{1,5}\s+[\w\s]+\b'
    addresses = re.findall(address_pattern, all_text)
    for address in addresses:
        # Split address into street number and name
        street_number, street_name = address.split(maxsplit=1)
        street_numbers.append(street_number)
        # Extract postal code and country from the last part of the address
        postal_code_country_pattern = r'(\d{4,})\s+([\w\s]+)$'
        match = re.search(postal_code_country_pattern, street_name)
        if match:
            postal_codes.append(match.group(1))
            countries.append(match.group(2))
        else:
            postal_codes.append('-')
            countries.append('-')
    print("Number of addresses found:", len(street_numbers))
    print("Number of postal codes found:", len(postal_codes))
    print("Number of countries found:", len(countries))
    
    # Assuming company name is not readily available in the webpage content
    company_names.extend(['-'] * len(email_addresses))
    
    # Assuming website URLs are not readily available in the webpage content
    websites.extend(['-'] * len(email_addresses))
    
    # Create a DataFrame to store the data
    data = {
        'Company Name': company_names,
        'Street & House number': street_numbers,
        'Postal Code': postal_codes,
        'Country': countries,
        'Email Address': email_addresses,
        'Phone Number': phone_numbers,
        'Fax Number': fax_numbers,
        'Website': websites
    }
    df = pd.DataFrame(data)
    
    return df

def extract_email_from_subpage(subpage_url):
    try:
        response = requests.get(subpage_url)
        response.raise_for_status()  
        soup = BeautifulSoup(response.text, 'html.parser')

        contact_block = soup.find('div', class_='cb-orte-item-adresse')

        if contact_block:
            email_element = contact_block.find('span', class_='data-emailencrypted').find_next('a', class_='value')
            email = email_element.get('href').split(':')[-1] if email_element else "Nicht gefunden"
            return email
        else:
            print("Kontaktblock nicht gefunden auf der Seite:", subpage_url)
            return "-"
    except Exception as e:
        print("Fehler beim Abrufen der Seite:", e)
        return "-"

def save_to_excel(df, filename='contact_data.xlsx'):
    # Function definition for saving DataFrame to Excel
    
    try:
        # Load existing Excel file
        wb = load_workbook(filename)
        # Select the active worksheet
        ws = wb.active
        # Find the next empty row in the worksheet
        next_row = ws.max_row + 1
        
        # Write the DataFrame to the Excel file starting from the next empty row
        for index, row in df.iterrows():
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=next_row+index, column=col_idx, value=value)
        
        # Save the updated workbook
        wb.save(filename)
        print("Data appended to", filename)
    
    except FileNotFoundError:
        # If the file doesn't exist, create a new Excel file
        df.to_excel(filename, index=False)
        print("New file created:", filename)

if __name__ == "__main__":
    urls = []
    while True:
        url = input("Enter a URL (or type 'done' to finish): ")
        if url.lower() == 'done':
            break
        urls.append(url)
    
    if not urls:
        print("No URLs provided. Exiting.")
    else:
        # Initialize an empty DataFrame to store all contact data
        all_contact_data = pd.DataFrame()
        
        # Iterate through each URL, scrape contact data, and concatenate to the DataFrame
        for url in urls:
            print("Scraping data from:", url)
            try:
                contact_df = scrape_contact_data(url)
                all_contact_data = pd.concat([all_contact_data, contact_df], ignore_index=True)
            except Exception as e:
                print(f"Error occurred while scraping data from {url}: {e}")
        
        # Specify the location where you want to save the Excel file
        save_location = "C:\\Users\\elard\\Desktop\\contact_data.xlsx"  # Replace with your desired location
        
        # Save all contact data to the specified location
        save_to_excel(all_contact_data, filename=save_location)
       
